#!/usr/bin/env python3
"""Inventory and profile likely scientific-figure inputs."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable


CATEGORIES = {
    "tabular": {".csv", ".tsv", ".xlsx", ".xls", ".parquet"},
    "structured": {".json", ".jsonl", ".yaml", ".yml"},
    "narrative": {".md", ".txt", ".tex", ".docx", ".pdf", ".rtf"},
    "raster": {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".webp", ".bmp"},
    "vector": {".svg", ".pdf", ".eps", ".ai", ".drawio"},
    "code": {".py", ".r", ".ipynb", ".m", ".jl"},
    "log": {".log", ".out"},
}

SKIP_DIRS = {".git", ".venv", "venv", "node_modules", "target", "dist", "build", "__pycache__"}
MAX_PROFILE_BYTES = 10_000_000
MAX_PROFILE_ROWS = 10_000
MAX_COLUMNS = 100
MAX_TEXT_PREVIEW = 2_000
MISSING = {"", "na", "n/a", "nan", "null", "none", "-"}
DATE_PATTERNS = ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S")


def categories_for(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    return [name for name, suffixes in CATEGORIES.items() if suffix in suffixes] or ["other"]


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip().lower() in MISSING


def to_number(value: Any) -> float | None:
    if isinstance(value, bool) or is_missing(value):
        return None
    try:
        number = float(str(value).strip().replace(",", ""))
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def is_date(value: Any) -> bool:
    text = str(value).strip()
    if not text or not re.search(r"[-/:]", text):
        return False
    for pattern in DATE_PATTERNS:
        try:
            datetime.strptime(text, pattern)
            return True
        except ValueError:
            continue
    return False


def profile_records(records: Iterable[dict[str, Any]], *, max_rows: int = MAX_PROFILE_ROWS) -> dict:
    rows: list[dict[str, Any]] = []
    columns: list[str] = []
    seen_columns: set[str] = set()
    for record in records:
        if not isinstance(record, dict):
            continue
        normalized = {str(key): value for key, value in record.items()}
        for key in normalized:
            if key not in seen_columns and len(columns) < MAX_COLUMNS:
                seen_columns.add(key)
                columns.append(key)
        rows.append(normalized)
        if len(rows) >= max_rows:
            break

    profiles = []
    for name in columns:
        values = [row.get(name) for row in rows]
        present = [value for value in values if not is_missing(value)]
        numeric = [number for value in present if (number := to_number(value)) is not None]
        date_count = sum(1 for value in present if is_date(value))
        unique_text = list(dict.fromkeys(str(value) for value in present))

        if present and len(numeric) == len(present):
            kind = "numeric"
        elif present and date_count == len(present):
            kind = "datetime"
        elif present and len(unique_text) <= min(30, max(2, len(present) // 2 + 1)):
            kind = "categorical"
        else:
            kind = "text"

        profile: dict[str, Any] = {
            "name": name,
            "type": kind,
            "missing": len(values) - len(present),
            "non_missing": len(present),
            "unique": len(unique_text),
            "sample_values": unique_text[:5],
        }
        if kind == "numeric" and numeric:
            profile.update({"min": min(numeric), "max": max(numeric)})
        profiles.append(profile)

    return {
        "row_count_profiled": len(rows),
        "truncated": len(rows) >= max_rows,
        "columns": profiles,
        "missing_cells": sum(column["missing"] for column in profiles),
        "data_candidate": bool(rows and profiles),
    }


def delimited_records(path: Path) -> Iterable[dict[str, Any]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        yield from csv.DictReader(handle, delimiter=delimiter)


def json_records(path: Path) -> Iterable[dict[str, Any]]:
    if path.suffix.lower() == ".jsonl":
        with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
            for line_number, line in enumerate(handle, start=1):
                if not line.strip():
                    continue
                value = json.loads(line)
                if not isinstance(value, dict):
                    raise ValueError(f"JSONL line {line_number} is not an object")
                yield value
        return

    value = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                yield item
    elif isinstance(value, dict):
        for candidate in ("records", "data", "results", "metrics", "items"):
            nested = value.get(candidate)
            if isinstance(nested, list):
                for item in nested:
                    if isinstance(item, dict):
                        yield item
                return
        yield value


def xlsx_records(path: Path) -> tuple[str, Iterable[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to profile .xlsx files") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else f"column_{index + 1}" for index, value in enumerate(next(iterator, []))]

    def records() -> Iterable[dict[str, Any]]:
        try:
            for values in iterator:
                yield dict(zip(headers, values))
        finally:
            workbook.close()

    return sheet.title, records()


def table_profile(path: Path) -> dict:
    if path.stat().st_size > MAX_PROFILE_BYTES:
        return {"profile_error": f"file exceeds {MAX_PROFILE_BYTES} byte profiling limit"}
    try:
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            profile = profile_records(delimited_records(path))
            profile["format"] = suffix.lstrip(".")
            return profile
        if suffix in {".json", ".jsonl"}:
            profile = profile_records(json_records(path))
            profile["format"] = suffix.lstrip(".")
            return profile
        if suffix == ".xlsx":
            sheet, records = xlsx_records(path)
            profile = profile_records(records)
            profile.update({"format": "xlsx", "sheet": sheet})
            return profile
        if suffix in {".xls", ".parquet"}:
            return {"profile_error": f"{suffix} profiling is not implemented; convert to CSV, JSON, or XLSX", "data_candidate": False}
    except (OSError, UnicodeError, csv.Error, json.JSONDecodeError, ValueError, RuntimeError) as exc:
        return {"profile_error": str(exc), "data_candidate": False}
    return {}


def narrative_preview(path: Path) -> str | None:
    suffix = path.suffix.lower()
    if path.stat().st_size > MAX_PROFILE_BYTES:
        return None
    try:
        if suffix in {".txt", ".md", ".tex"}:
            return path.read_text(encoding="utf-8-sig", errors="replace")[:MAX_TEXT_PREVIEW].strip()
        if suffix == ".docx":
            with zipfile.ZipFile(path) as archive:
                xml = archive.read("word/document.xml")
            root = ET.fromstring(xml)
            text = " ".join(node.text or "" for node in root.iter() if node.tag.rsplit("}", 1)[-1] == "t")
            return text[:MAX_TEXT_PREVIEW].strip()
        if suffix == ".pdf":
            try:
                from pypdf import PdfReader  # type: ignore
            except ImportError:
                return None
            reader = PdfReader(path)
            text = "\n".join((page.extract_text() or "") for page in reader.pages[:5])
            return text[:MAX_TEXT_PREVIEW].strip()
    except (OSError, KeyError, ET.ParseError, zipfile.BadZipFile):
        return None
    return None


def inventory(root: Path) -> dict:
    files = []
    counts: Counter[str] = Counter()
    for path in sorted(root.rglob("*")):
        if not path.is_file() or any(part in SKIP_DIRS for part in path.relative_to(root).parts):
            continue
        cats = categories_for(path)
        counts.update(cats)
        entry: dict[str, Any] = {
            "path": path.relative_to(root).as_posix(),
            "size_bytes": path.stat().st_size,
            "categories": cats,
        }
        profile = table_profile(path)
        if profile:
            entry["table_profile"] = profile
        preview = narrative_preview(path)
        if preview:
            entry["text_preview"] = preview
        files.append(entry)
    return {
        "schema_version": "1.1",
        "root": str(root.resolve()),
        "file_count": len(files),
        "category_counts": dict(sorted(counts.items())),
        "data_file_count": sum(1 for item in files if item.get("table_profile", {}).get("data_candidate")),
        "files": files,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("root", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    if not args.root.is_dir():
        parser.error(f"input root is not a directory: {args.root}")
    result = inventory(args.root)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Inventoried {result['file_count']} files ({result['data_file_count']} data candidates) -> {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
