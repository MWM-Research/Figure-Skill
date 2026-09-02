from __future__ import annotations
import csv, importlib.util, json, math
from pathlib import Path
from typing import Any
from .structural import check, sha256

def load_statistics_core():
    path = Path(__file__).resolve().parent.parent / "backends" / "statistics_core.py"
    spec = importlib.util.spec_from_file_location("qa_statistics_core", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load statistics core: {path}")
    module = importlib.util.module_from_spec(spec); spec.loader.exec_module(module); return module


STATS = load_statistics_core()

def equal_value(raw: Any, expected: Any) -> bool:
    try:
        return math.isclose(float(str(raw).replace(",", "")), float(expected), rel_tol=1e-9, abs_tol=1e-12)
    except (TypeError, ValueError):
        return str(raw) == str(expected)


def source_rows(path: Path) -> dict[int, dict[str, Any]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            return {index: row for index, row in enumerate(csv.DictReader(handle, delimiter=delimiter), start=2)}
    if suffix == ".jsonl":
        result = {}
        with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
            for index, line in enumerate(handle, start=1):
                if line.strip():
                    value = json.loads(line)
                    if isinstance(value, dict):
                        result[index] = value
        return result
    if suffix == ".json":
        value = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(value, dict):
            value = next((value[key] for key in ("records", "data", "results", "metrics", "items") if isinstance(value.get(key), list)), [value])
        return {index: item for index, item in enumerate(value, start=1) if isinstance(item, dict)} if isinstance(value, list) else {}
    if suffix == ".xlsx":
        from openpyxl import load_workbook  # type: ignore
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else f"column_{index + 1}" for index, value in enumerate(next(iterator, []))]
        try:
            return {index: dict(zip(headers, values)) for index, values in enumerate(iterator, start=2)}
        finally:
            workbook.close()
    raise ValueError(f"unsupported provenance source format: {suffix}")


def values_close(left: Any, right: Any) -> bool:
    if isinstance(left, dict) and isinstance(right, dict):
        return left.keys() == right.keys() and all(values_close(left[key], right[key]) for key in left)
    if isinstance(left, list) and isinstance(right, list):
        return len(left) == len(right) and all(values_close(a, b) for a, b in zip(left, right))
    return equal_value(left, right)


def plan_panel_index(plan: dict | None) -> dict[tuple[str, str | None], dict]:
    result = {}
    for panel in (plan or {}).get("panels", []):
        panel_id = str(panel.get("id"))
        result[(panel_id, None)] = panel
        for subplot in panel.get("subplots", []):
            result[(panel_id, str(subplot.get("id")))] = subplot
    return result


def recompute_derived(mark: dict, panel: dict, rows: dict[int, dict[str, Any]]) -> tuple[bool, str]:
    derived = mark.get("derived")
    if not isinstance(derived, dict): return True, ""
    operation = derived.get("operation")
    source_row_numbers = [int(value) for value in derived.get("source_rows", mark.get("source_rows", []))]
    selected = [rows[number] for number in source_row_numbers if number in rows]
    if len(selected) != len(source_row_numbers): return False, "derived source row missing"
    try:
        if operation == "box-summary":
            column = str(panel["y"]); expected = STATS.box_summary([float(row[column]) for row in selected], source_row_numbers)
        elif operation == "kde":
            column = str(panel.get("value") or panel.get("y")); expected = STATS.kde([float(row[column]) for row in selected], source_row_numbers, bandwidth=derived["parameters"]["bandwidth"])
        elif operation == "histogram":
            column = str(panel.get("value") or panel.get("x")); expected = STATS.histogram([float(row[column]) for row in selected], source_row_numbers, derived["parameters"])
        elif operation == "confusion-count":
            actual, predicted = str(panel["actual"]), str(panel["predicted"]); expected = STATS.confusion_matrix([str(row[actual]) for row in selected], [str(row[predicted]) for row in selected], source_row_numbers, derived["parameters"]["normalization"])
        elif operation in {"roc", "pr"}:
            label, score = str(panel["label"]), str(panel["score"]); expected = STATS.binary_curve([row[label] for row in selected], [float(row[score]) for row in selected], source_row_numbers, derived["parameters"]["positive_label"], operation, bool(derived["parameters"].get("compute_auc")))
        else:
            return False, f"unsupported derived operation {operation}"
        return values_close(expected, derived), "" if values_close(expected, derived) else f"derived {operation} mismatch"
    except (KeyError, TypeError, ValueError) as exc:
        return False, str(exc)


def verify_data_provenance(path: Path, plan: dict | None = None) -> list[dict]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return [check("provenance-readable", "fail", file=str(path), detail=str(exc))]
    checks = [check("provenance-readable", "pass", file=str(path))]
    panel_lookup = plan_panel_index(plan)
    for panel in data.get("panels", []):
        panel_id = panel.get("panel")
        sources = {str(panel.get("source_file", "")): panel.get("source_sha256")} if panel.get("source_file") else {str(item.get("file")): item.get("sha256") for item in panel.get("sources", [])}
        if not sources:
            checks.append(check("provenance-source-exists", "fail", panel=panel_id, source="")); continue
        source_rows_cache = {}
        for source_text, expected_hash in sources.items():
            source = Path(source_text)
            checks.append(check("provenance-source-exists", "pass" if source.is_file() else "fail", panel=panel_id, source=str(source)))
            if source.is_file():
                checks.append(check("provenance-source-hash", "pass" if expected_hash and sha256(source) == expected_hash else "fail", panel=panel_id, source=str(source)))
                source_rows_cache[str(source)] = source_rows(source)
        try:
            mismatches = []
            derived_mismatches = []
            for mark in panel.get("marks", []):
                source_text = str(mark.get("source_file") or panel.get("source_file"))
                rows = source_rows_cache.get(source_text, {})
                row_numbers = [int(mark["source_row"])] if mark.get("source_row") is not None else [int(value) for value in mark.get("source_rows", [])]
                if any(number not in rows for number in row_numbers): mismatches.append(f"missing source row(s) {row_numbers}")
                if mark.get("source_row") is not None:
                    row_number = int(mark["source_row"]); row = rows.get(row_number, {})
                    for axis in ("x", "y", "group", "value", "error", "x2"):
                        value = mark.get(axis, {})
                        if not value or value.get("column") is None: continue
                        column = value.get("column")
                        if column not in row or not equal_value(row.get(column), value.get("value")): mismatches.append(f"row {row_number} column {column}")
                    uncertainty = mark.get("uncertainty", {})
                    for label, column in uncertainty.get("columns", {}).items():
                        expected = uncertainty.get("values", {}).get(label)
                        if column not in row or not equal_value(row.get(column), expected): mismatches.append(f"row {row_number} column {column}")
                plan_panel = panel_lookup.get((str(panel_id), str(mark.get("subplot")) if mark.get("subplot") is not None else None))
                if mark.get("derived") and plan_panel:
                    valid, detail = recompute_derived(mark, plan_panel, rows)
                    if not valid: derived_mismatches.append(detail)
            checks.append(check(
                "provenance-mark-values", "fail" if mismatches else "pass", panel=panel_id,
                detail=", ".join(mismatches) if mismatches else f"verified {len(panel.get('marks', []))} marks",
            ))
            checks.append(check("provenance-derived-values", "fail" if derived_mismatches else "pass", panel=panel_id, detail=", ".join(dict.fromkeys(derived_mismatches)) if derived_mismatches else "verified"))
        except (OSError, csv.Error, ValueError, TypeError) as exc:
            checks.append(check("provenance-mark-values", "fail", panel=panel_id, detail=str(exc)))
    return checks
