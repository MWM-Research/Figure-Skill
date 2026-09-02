from __future__ import annotations
import csv, hashlib, json
from pathlib import Path
from typing import Any

def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_records(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            return [dict(row, __source_row__=index) for index, row in enumerate(csv.DictReader(handle, delimiter=delimiter), start=2)]
    if suffix == ".jsonl":
        records = []
        with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
            for index, line in enumerate(handle, start=1):
                if line.strip():
                    value = json.loads(line)
                    if not isinstance(value, dict):
                        raise ValueError(f"JSONL line {index} is not an object")
                    records.append(dict(value, __source_row__=index))
        return records
    if suffix == ".json":
        value = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(value, dict):
            value = next((value[key] for key in ("records", "data", "results", "metrics", "items") if isinstance(value.get(key), list)), [value])
        if not isinstance(value, list):
            raise ValueError("JSON input must be an object or a list of objects")
        return [dict(item, __source_row__=index) for index, item in enumerate(value, start=1) if isinstance(item, dict)]
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to render .xlsx data") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else f"column_{index + 1}" for index, value in enumerate(next(iterator, []))]
        try:
            return [dict(zip(headers, values), __source_row__=index) for index, values in enumerate(iterator, start=2)]
        finally:
            workbook.close()
    raise ValueError(f"unsupported data format: {path.suffix}")

def panel_records(panel: dict, input_root: Path) -> tuple[Path, list[dict[str, Any]]]:
    sources = panel.get("source_files") or []
    if len(sources) != 1:
        raise ValueError(f"panel {panel.get('id')} must reference exactly one data source")
    source = (input_root / sources[0]).resolve()
    try:
        source.relative_to(input_root.resolve())
    except ValueError as exc:
        raise ValueError(f"source escapes input root: {sources[0]}") from exc
    if not source.is_file():
        raise FileNotFoundError(f"data source not found: {source}")
    return source, read_records(source)

