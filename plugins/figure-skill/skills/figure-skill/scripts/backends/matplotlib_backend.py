#!/usr/bin/env python3
"""Render evidence-linked data panels from a scientific figure plan."""

from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import json
import math
import shutil
from pathlib import Path
from typing import Any, Iterable


def load_statistics_core():
    path = Path(__file__).with_name("statistics_core.py")
    spec = importlib.util.spec_from_file_location("figure_statistics_core", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load statistics core: {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


STATS = load_statistics_core()


def require_matplotlib():
    try:
        import matplotlib.pyplot as plt  # type: ignore
    except ImportError as exc:
        raise SystemExit("matplotlib is required; install the dependencies listed in this Skill's requirements.txt") from exc
    return plt


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_records(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            return [dict(row, __source_row__=index) for index, row in enumerate(csv.DictReader(handle, delimiter=delimiter), start=2)]
    if suffix == ".jsonl":
        records = []
        with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
            for index, line in enumerate(handle, start=1):
                if line.strip():
                    value = json.loads(line)
                    if not isinstance(value, dict):
                        raise ValueError(f"JSONL line {index} is not an object")
                    records.append(dict(value, __source_row__=index))
        return records
    if suffix == ".json":
        value = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(value, dict):
            value = next((value[key] for key in ("records", "data", "results", "metrics", "items") if isinstance(value.get(key), list)), [value])
        if not isinstance(value, list):
            raise ValueError("JSON input must be an object or a list of objects")
        return [dict(item, __source_row__=index) for index, item in enumerate(value, start=1) if isinstance(item, dict)]
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to render .xlsx data") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else f"column_{index + 1}" for index, value in enumerate(next(iterator, []))]
        try:
            return [dict(zip(headers, values), __source_row__=index) for index, values in enumerate(iterator, start=2)]
        finally:
            workbook.close()
    raise ValueError(f"unsupported data format: {path.suffix}")


def finite_number(value: Any, *, column: str, row: int) -> float:
    try:
        number = float(str(value).strip().replace(",", ""))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"non-numeric value in column {column!r}, source row {row}: {value!r}") from exc
    if not math.isfinite(number):
        raise ValueError(f"non-finite value in column {column!r}, source row {row}: {value!r}")
    return number


def error_number(value: Any, *, column: str, row: int) -> float:
    number = finite_number(value, column=column, row=row)
    if number < 0:
        raise ValueError(f"negative uncertainty in column {column!r}, source row {row}: {value!r}")
    return number


def uncertainty_config(panel: dict) -> dict[str, Any] | None:
    configured = panel.get("uncertainty")
    legacy = panel.get("error")
    if configured and legacy:
        raise ValueError("use either uncertainty or legacy error, not both")
    if configured:
        if not isinstance(configured, dict):
            raise ValueError("uncertainty must be an object")
        return configured
    return {"mode": "symmetric-delta", "error_column": legacy} if legacy else None


def record_uncertainty(record: dict[str, Any], panel: dict, y_value: float) -> dict[str, Any] | None:
    config = uncertainty_config(panel)
    if not config:
        return None
    mode = config.get("mode")
    row = int(record["__source_row__"])
    if mode == "symmetric-delta":
        column = config.get("error_column")
        if not column:
            raise ValueError("symmetric-delta uncertainty requires error_column")
        value = error_number(record.get(column), column=str(column), row=row)
        return {"mode": mode, "lower": value, "upper": value, "columns": {"error": str(column)}, "values": {"error": value}}
    lower_column, upper_column = config.get("lower_column"), config.get("upper_column")
    if not lower_column or not upper_column:
        raise ValueError(f"{mode} uncertainty requires lower_column and upper_column")
    lower_value = finite_number(record.get(lower_column), column=str(lower_column), row=row)
    upper_value = finite_number(record.get(upper_column), column=str(upper_column), row=row)
    if mode == "asymmetric-delta":
        if lower_value < 0 or upper_value < 0:
            raise ValueError("asymmetric uncertainty deltas must be non-negative")
        lower_delta, upper_delta = lower_value, upper_value
    elif mode == "bounds":
        if not lower_value <= y_value <= upper_value:
            raise ValueError(f"uncertainty bounds must satisfy lower <= y <= upper at source row {row}")
        lower_delta, upper_delta = y_value - lower_value, upper_value - y_value
    else:
        raise ValueError("uncertainty mode must be symmetric-delta, asymmetric-delta, or bounds")
    return {
        "mode": mode, "lower": lower_delta, "upper": upper_delta,
        "columns": {"lower": str(lower_column), "upper": str(upper_column)},
        "values": {"lower": lower_value, "upper": upper_value},
    }


def panel_records(panel: dict, input_root: Path) -> tuple[Path, list[dict[str, Any]]]:
    sources = panel.get("source_files") or []
    if len(sources) != 1:
        raise ValueError(f"panel {panel.get('id')} must reference exactly one data source")
    source = (input_root / sources[0]).resolve()
    try:
        source.relative_to(input_root.resolve())
    except ValueError as exc:
        raise ValueError(f"source escapes input root: {sources[0]}") from exc
    if not source.is_file():
        raise FileNotFoundError(f"data source not found: {source}")
    return source, read_records(source)


def plot_bar(ax, panel: dict, records: list[dict[str, Any]]) -> list[dict]:
    x_name, y_name = panel["x"], panel["y"]
    group_name = panel.get("group")
    labels = [str(record.get(x_name, "")) for record in records]
    if not labels or any(not label for label in labels):
        raise ValueError(f"panel {panel['id']} contains an empty {x_name!r} category")
    if not group_name and len(set(labels)) != len(labels):
        raise ValueError(f"panel {panel['id']} has duplicate {x_name!r} categories; define an aggregation before plotting")
    values = [finite_number(record.get(y_name), column=y_name, row=int(record["__source_row__"])) for record in records]
    uncertainties = [record_uncertainty(record, panel, value) for record, value in zip(records, values)]
    marks = []
    if group_name:
        groups = [str(record.get(group_name, "")) for record in records]
        if any(not group for group in groups):
            raise ValueError(f"panel {panel['id']} contains an empty {group_name!r} group")
        pairs = list(zip(labels, groups))
        if len(set(pairs)) != len(pairs):
            raise ValueError(f"panel {panel['id']} has duplicate ({x_name}, {group_name}) pairs; define an aggregation")
        x_labels = list(dict.fromkeys(labels))
        group_labels = list(dict.fromkeys(groups))
        lookup = {pair: (value, uncertainty, record) for pair, value, uncertainty, record in zip(pairs, values, uncertainties, records)}
        width = 0.8 / len(group_labels)
        palette = ["#b8b8b8", "#3f6f8f", "#7f7f7f", "#7b9e87", "#8b6f8f"]
        for group_index, group in enumerate(group_labels):
            positions = [index - 0.4 + width / 2 + group_index * width for index in range(len(x_labels))]
            group_values = [lookup.get((label, group), (math.nan, None, None))[0] for label in x_labels]
            group_uncertainties = [lookup.get((label, group), (math.nan, None, None))[1] for label in x_labels]
            group_errors = [
                [item["lower"] if item else math.nan for item in group_uncertainties],
                [item["upper"] if item else math.nan for item in group_uncertainties],
            ]
            bars = ax.bar(
                positions, group_values, width=width, label=group,
                color=palette[group_index % len(palette)], edgecolor="#222222", linewidth=0.8,
                yerr=group_errors if uncertainty_config(panel) else None,
                capsize=3 if uncertainty_config(panel) else 0,
            )
            for bar, label, value, uncertainty in zip(bars, x_labels, group_values, group_uncertainties):
                if not math.isnan(value):
                    ax.text(bar.get_x() + bar.get_width() / 2, value, f"{value:g}", ha="center", va="bottom", fontsize=7.5)
                    record = lookup[(label, group)][2]
                    mark = {"source_row": int(record["__source_row__"]), "x": label, "group": group, "y": value}
                    if uncertainty is not None:
                        mark["uncertainty"] = uncertainty
                    marks.append(mark)
        ax.set_xticks(range(len(x_labels)), x_labels)
        ax.legend(title=group_name, frameon=False)
    else:
        colors = ["#b8b8b8"] * len(values)
        if colors:
            colors[-1] = "#3f6f8f"
        bars = ax.bar(
            labels, values, width=0.62, color=colors, edgecolor="#222222", linewidth=0.9,
            yerr=[
                [item["lower"] if item else 0.0 for item in uncertainties],
                [item["upper"] if item else 0.0 for item in uncertainties],
            ] if uncertainty_config(panel) else None,
            capsize=3 if uncertainty_config(panel) else 0,
        )
        for bar, value in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, value, f"{value:g}", ha="center", va="bottom", fontsize=8.5)
        marks = []
        for record, label, value, uncertainty in zip(records, labels, values, uncertainties):
            mark = {"source_row": int(record["__source_row__"]), "x": label, "y": value}
            if uncertainty is not None:
                mark["uncertainty"] = uncertainty
            marks.append(mark)
    lower_values = [value - (item["lower"] if item else 0.0) for value, item in zip(values, uncertainties)]
    upper_values = [value + (item["upper"] if item else 0.0) for value, item in zip(values, uncertainties)]
    lower = min(0.0, min(lower_values) * 1.16)
    upper = max(0.0, max(upper_values) * 1.16)
    if lower == upper:
        upper = lower + 1.0
    ax.set_ylim(lower, upper)
    ax.set_xlabel(x_name)
    ax.set_ylabel(y_name if not panel.get("unit") else f"{y_name} ({panel['unit']})")
    ax.grid(axis="y", color="#dddddd", linewidth=0.7)
    ax.set_axisbelow(True)
    ax.spines[["top", "right"]].set_visible(False)
    return marks


def plot_xy(ax, panel: dict, records: list[dict[str, Any]], *, scatter: bool) -> list[dict]:
    x_name, y_name = panel["x"], panel["y"]
    group_name = panel.get("group")
    grouped: dict[str, list[tuple[float, float, dict[str, Any] | None, int]]] = {}
    for record in records:
        row = int(record["__source_row__"])
        x_value = finite_number(record.get(x_name), column=x_name, row=row)
        y_value = finite_number(record.get(y_name), column=y_name, row=row)
        uncertainty = record_uncertainty(record, panel, y_value)
        group = str(record.get(group_name, "")) if group_name else ""
        if group_name and not group:
            raise ValueError(f"panel {panel['id']} contains an empty {group_name!r} group")
        grouped.setdefault(group, []).append((x_value, y_value, uncertainty, row))
    palette = ["#3f6f8f", "#7b9e87", "#8b6f8f", "#c17c4f", "#6b7280"]
    marks = []
    for group_index, (group, points) in enumerate(grouped.items()):
        points.sort(key=lambda point: point[0])
        xs = [point[0] for point in points]
        if len(set(xs)) != len(xs):
            scope = f" within group {group!r}" if group_name else ""
            raise ValueError(f"panel {panel['id']} has duplicate {x_name!r} values{scope}; define an aggregation")
        ys = [point[1] for point in points]
        uncertainties = [point[2] for point in points]
        color = palette[group_index % len(palette)]
        if uncertainty_config(panel):
            ax.errorbar(
                xs, ys, yerr=[
                    [item["lower"] if item else 0.0 for item in uncertainties],
                    [item["upper"] if item else 0.0 for item in uncertainties],
                ], fmt="o",
                linestyle="none" if scatter else "-", color=color, ecolor=color,
                linewidth=1.6, markersize=4.5, capsize=3, label=group if group_name else None,
            )
        elif scatter:
            ax.scatter(xs, ys, s=34, color=color, edgecolor="#222222", linewidth=0.6, label=group if group_name else None)
        else:
            ax.plot(xs, ys, marker="o", color=color, linewidth=1.6, markersize=4.5, label=group if group_name else None)
        for x_value, y_value, uncertainty, row in points:
            mark = {"source_row": row, "x": x_value, "y": y_value}
            if group_name:
                mark["group"] = group
            if uncertainty is not None:
                mark["uncertainty"] = uncertainty
            marks.append(mark)
    ax.set_xlabel(x_name)
    ax.set_ylabel(y_name if not panel.get("unit") else f"{y_name} ({panel['unit']})")
    ax.grid(color="#dddddd", linewidth=0.7)
    ax.set_axisbelow(True)
    ax.spines[["top", "right"]].set_visible(False)
    if group_name:
        ax.legend(title=group_name, frameon=False)
    return marks


def plot_heatmap(ax, fig, panel: dict, records: list[dict[str, Any]]) -> list[dict]:
    import numpy as np  # type: ignore

    x_name, y_name, value_name = panel["x"], panel["y"], panel["value"]
    x_labels = list(dict.fromkeys(str(record.get(x_name, "")) for record in records))
    y_labels = list(dict.fromkeys(str(record.get(y_name, "")) for record in records))
    if not x_labels or not y_labels or any(not label for label in x_labels + y_labels):
        raise ValueError(f"panel {panel['id']} contains an empty heatmap coordinate")
    x_index = {label: index for index, label in enumerate(x_labels)}
    y_index = {label: index for index, label in enumerate(y_labels)}
    matrix = np.full((len(y_labels), len(x_labels)), np.nan, dtype=float)
    seen: set[tuple[str, str]] = set()
    marks = []
    for record in records:
        row = int(record["__source_row__"])
        x_label = str(record.get(x_name, ""))
        y_label = str(record.get(y_name, ""))
        coordinate = (x_label, y_label)
        if coordinate in seen:
            raise ValueError(f"panel {panel['id']} has duplicate heatmap coordinate {coordinate}; define an aggregation")
        seen.add(coordinate)
        value = finite_number(record.get(value_name), column=value_name, row=row)
        matrix[y_index[y_label], x_index[x_label]] = value
        marks.append({"source_row": row, "x": x_label, "y": y_label, "value": value})
    missing = int(np.isnan(matrix).sum())
    if missing:
        raise ValueError(f"panel {panel['id']} heatmap grid is incomplete: {missing} coordinate(s) missing")
    mesh = ax.pcolormesh(
        np.arange(len(x_labels) + 1), np.arange(len(y_labels) + 1), matrix,
        cmap=str(panel.get("colormap") or "viridis"), shading="flat",
        edgecolors="#FFFFFF", linewidth=0.6, rasterized=False,
    )
    ax.set_xticks(np.arange(len(x_labels)) + 0.5, x_labels)
    ax.set_yticks(np.arange(len(y_labels)) + 0.5, y_labels)
    ax.invert_yaxis()
    ax.set_xlabel(x_name)
    ax.set_ylabel(y_name)
    colorbar = fig.colorbar(mesh, ax=ax, pad=0.025)
    if colorbar.solids is not None:
        colorbar.solids.set_rasterized(False)
    colorbar.set_label(value_name if not panel.get("unit") else f"{value_name} ({panel['unit']})")
    if panel.get("annotate_values", len(marks) <= 64):
        threshold = (float(matrix.min()) + float(matrix.max())) / 2
        for mark in marks:
            ax.text(
                x_index[mark["x"]] + 0.5, y_index[mark["y"]] + 0.5, f"{mark['value']:g}",
                ha="center", va="center", fontsize=7.5,
                color="#FFFFFF" if mark["value"] > threshold else "#111827",
            )
    return marks


def calculation(panel: dict, operation: str) -> dict[str, Any]:
    value = panel.get("calculation") or {"mode": "precomputed", "operation": operation, "parameters": {}}
    if not isinstance(value, dict) or value.get("mode") not in {"precomputed", "raw"}:
        raise ValueError("calculation.mode must be precomputed or raw")
    if value.get("operation") not in {None, operation}:
        raise ValueError(f"visual form requires calculation.operation={operation}")
    parameters = value.get("parameters", {})
    if not isinstance(parameters, dict):
        raise ValueError("calculation.parameters must be an object")
    return {"mode": value["mode"], "operation": operation, "parameters": parameters}


def grouped_samples(panel: dict, records: list[dict[str, Any]], value_name: str) -> dict[str, tuple[list[float], list[int]]]:
    group_name = panel.get("group") or panel.get("x")
    grouped: dict[str, tuple[list[float], list[int]]] = {}
    for record in records:
        row = int(record["__source_row__"])
        group = str(record.get(group_name, "All")) if group_name else "All"
        if not group:
            raise ValueError(f"panel {panel['id']} contains an empty group")
        values, rows = grouped.setdefault(group, ([], []))
        values.append(finite_number(record.get(value_name), column=value_name, row=row))
        rows.append(row)
    return grouped


def plot_box(ax, panel: dict, records: list[dict[str, Any]]) -> list[dict]:
    config = calculation(panel, "box-summary")
    labels, stats, marks = [], [], []
    if config["mode"] == "raw":
        grouped = grouped_samples(panel, records, str(panel["y"]))
        for label, (values, rows) in grouped.items():
            summary = STATS.box_summary(values, rows)
            labels.append(label)
            stats.append({
                "label": label, "q1": summary["q1"], "med": summary["median"], "q3": summary["q3"],
                "whislo": summary["whisker_low"], "whishi": summary["whisker_high"],
                "fliers": [item["value"] for item in summary["outliers"]],
            })
            marks.append({"group": label, "source_rows": rows, "derived": summary})
    else:
        x_name = str(panel["x"])
        fields = config["parameters"].get("columns", {
            "q1": "q1", "median": "median", "q3": "q3", "whisker_low": "whisker_low", "whisker_high": "whisker_high",
        })
        for record in records:
            row, label = int(record["__source_row__"]), str(record.get(x_name, ""))
            values = {key: finite_number(record.get(column), column=column, row=row) for key, column in fields.items()}
            if not values["whisker_low"] <= values["q1"] <= values["median"] <= values["q3"] <= values["whisker_high"]:
                raise ValueError(f"invalid precomputed box ordering at source row {row}")
            labels.append(label)
            stats.append({"label": label, "q1": values["q1"], "med": values["median"], "q3": values["q3"], "whislo": values["whisker_low"], "whishi": values["whisker_high"], "fliers": []})
            marks.append({"group": label, "source_row": row, "precomputed": {"columns": fields, "values": values}})
    ax.bxp(stats, showfliers=True, patch_artist=True, boxprops={"facecolor": "#b8c7d1", "edgecolor": "#222222"}, medianprops={"color": "#1d4ed8", "linewidth": 1.6})
    ax.set_xticks(range(1, len(labels) + 1), labels)
    ax.set_xlabel(str(panel.get("x") or panel.get("group") or "group"))
    ax.set_ylabel(str(panel["y"]))
    ax.grid(axis="y", color="#dddddd", linewidth=0.7)
    return marks


def plot_density_or_violin(ax, panel: dict, records: list[dict[str, Any]], *, violin: bool) -> list[dict]:
    config = calculation(panel, "kde")
    marks = []
    if config["mode"] == "raw":
        value_name = str(panel.get("value") or panel.get("y"))
        grouped = grouped_samples(panel, records, value_name)
        bandwidth = config["parameters"].get("bandwidth")
        if bandwidth is None:
            raise ValueError("raw KDE requires an explicit bandwidth: 'scott' or a positive number")
        for index, (label, (values, rows)) in enumerate(grouped.items(), start=1):
            derived = STATS.kde(values, rows, bandwidth=bandwidth)
            grid, density = derived["grid"], derived["density"]
            if violin:
                maximum = max(density) or 1.0
                scaled = [value / maximum * 0.38 for value in density]
                ax.fill_betweenx(grid, [index - value for value in scaled], [index + value for value in scaled], alpha=0.65, color="#7b9e87", edgecolor="#222222")
            else:
                ax.plot(grid, density, linewidth=1.6, label=label)
            marks.append({"group": label, "source_rows": rows, "derived": derived})
        if violin:
            ax.set_xticks(range(1, len(grouped) + 1), list(grouped))
            ax.set_ylabel(value_name)
        else:
            ax.set_xlabel(value_name)
            ax.set_ylabel("density")
            if len(grouped) > 1:
                ax.legend(frameon=False, title=str(panel.get("group") or panel.get("x") or "group"))
    else:
        x_name, y_name = str(panel["x"]), str(panel["y"])
        density_name = str(panel.get("value") or config["parameters"].get("density_column") or "density")
        group_name = panel.get("group")
        grouped_rows: dict[str, list[dict[str, Any]]] = {}
        for record in records:
            grouped_rows.setdefault(str(record.get(group_name, "All")) if group_name else "All", []).append(record)
        for index, (label, items) in enumerate(grouped_rows.items(), start=1):
            points = sorted((finite_number(item.get(x_name), column=x_name, row=int(item["__source_row__"])), finite_number(item.get(density_name), column=density_name, row=int(item["__source_row__"])), int(item["__source_row__"])) for item in items)
            grid, density = [item[0] for item in points], [item[1] for item in points]
            if violin:
                maximum = max(density) or 1.0
                scaled = [value / maximum * 0.38 for value in density]
                ax.fill_betweenx(grid, [index - value for value in scaled], [index + value for value in scaled], alpha=0.65, color="#7b9e87", edgecolor="#222222")
            else:
                ax.plot(grid, density, label=label)
            marks.extend({"source_row": row, "x": x_value, "y": density_value, "group": label} for x_value, density_value, row in points)
        if violin:
            ax.set_xticks(range(1, len(grouped_rows) + 1), list(grouped_rows))
            ax.set_ylabel(y_name)
        elif len(grouped_rows) > 1:
            ax.legend(frameon=False)
    ax.grid(color="#dddddd", linewidth=0.7)
    return marks


def plot_histogram(ax, panel: dict, records: list[dict[str, Any]]) -> list[dict]:
    config = calculation(panel, "histogram")
    marks = []
    if config["mode"] == "raw":
        value_name = str(panel.get("value") or panel.get("x"))
        grouped = grouped_samples(panel, records, value_name)
        for label, (values, rows) in grouped.items():
            derived = STATS.histogram(values, rows, config["parameters"])
            edges, counts = derived["edges"], derived["counts"]
            ax.stairs(counts, edges, fill=len(grouped) == 1, alpha=0.45, linewidth=1.5, label=label)
            marks.append({"group": label, "source_rows": rows, "derived": derived})
        if len(grouped) > 1:
            ax.legend(frameon=False)
        ax.set_xlabel(value_name)
        ax.set_ylabel("density" if config["parameters"].get("density") else "count")
    else:
        left_name, right_name, height_name = str(panel["x"]), str(panel.get("x2") or "bin_right"), str(panel["y"])
        for record in records:
            row = int(record["__source_row__"])
            left = finite_number(record.get(left_name), column=left_name, row=row)
            right = finite_number(record.get(right_name), column=right_name, row=row)
            height = finite_number(record.get(height_name), column=height_name, row=row)
            if right <= left:
                raise ValueError(f"histogram bin right must exceed left at source row {row}")
            ax.bar((left + right) / 2, height, width=right - left, align="center", color="#8da9b8", edgecolor="#222222")
            marks.append({"source_row": row, "x": left, "x2": right, "y": height})
        ax.set_xlabel(left_name)
        ax.set_ylabel(height_name)
    ax.grid(axis="y", color="#dddddd", linewidth=0.7)
    return marks


def plot_confusion(ax, fig, panel: dict, records: list[dict[str, Any]]) -> list[dict]:
    import numpy as np  # type: ignore
    config = calculation(panel, "confusion-count")
    if config["mode"] == "raw":
        actual_name, predicted_name = str(panel["actual"]), str(panel["predicted"])
        rows = [int(record["__source_row__"]) for record in records]
        derived = STATS.confusion_matrix(
            [str(record.get(actual_name, "")) for record in records],
            [str(record.get(predicted_name, "")) for record in records], rows,
            str(config["parameters"].get("normalization", "none")),
        )
        labels, matrix = derived["labels"], np.asarray(derived["matrix"], dtype=float)
        marks = []
        for y, actual in enumerate(labels):
            for x, predicted in enumerate(labels):
                marks.append({"x": predicted, "y": actual, "value": float(matrix[y, x]), "source_rows": derived["cell_source_rows"][y][x], "derived": derived})
    else:
        x_name, y_name, value_name = str(panel["x"]), str(panel["y"]), str(panel["value"])
        x_labels = list(dict.fromkeys(str(record.get(x_name, "")) for record in records))
        y_labels = list(dict.fromkeys(str(record.get(y_name, "")) for record in records))
        matrix = np.full((len(y_labels), len(x_labels)), np.nan)
        marks = []
        for record in records:
            row = int(record["__source_row__"]); x = str(record.get(x_name, "")); y = str(record.get(y_name, ""))
            if not math.isnan(matrix[y_labels.index(y), x_labels.index(x)]):
                raise ValueError(f"duplicate confusion coordinate {(x, y)}")
            value = finite_number(record.get(value_name), column=value_name, row=row)
            matrix[y_labels.index(y), x_labels.index(x)] = value
            marks.append({"source_row": row, "x": x, "y": y, "value": value})
        if np.isnan(matrix).any():
            raise ValueError("precomputed confusion matrix must be rectangular and complete")
        labels = x_labels
        if x_labels != y_labels:
            raise ValueError("precomputed confusion matrix must use the same ordered actual/predicted labels")
    mesh = ax.pcolormesh(np.arange(len(labels) + 1), np.arange(len(labels) + 1), matrix, cmap=str(panel.get("colormap") or "Blues"), edgecolors="#ffffff", linewidth=0.6, rasterized=False)
    ax.set_xticks(np.arange(len(labels)) + 0.5, labels, rotation=30, ha="right")
    ax.set_yticks(np.arange(len(labels)) + 0.5, labels); ax.invert_yaxis(); ax.set_xlabel("Predicted"); ax.set_ylabel("Actual")
    colorbar = fig.colorbar(mesh, ax=ax, pad=0.025)
    if colorbar.solids is not None: colorbar.solids.set_rasterized(False)
    return marks


def plot_curve(ax, panel: dict, records: list[dict[str, Any]], *, curve: str) -> list[dict]:
    config = calculation(panel, curve)
    marks = []
    if config["mode"] == "raw":
        label_name, score_name = str(panel["label"]), str(panel["score"])
        positive_label = config["parameters"].get("positive_label")
        if positive_label is None:
            raise ValueError("raw ROC/PR requires positive_label")
        group_name = panel.get("group")
        grouped: dict[str, list[dict[str, Any]]] = {}
        for record in records:
            grouped.setdefault(str(record.get(group_name, "All")) if group_name else "All", []).append(record)
        for group, items in grouped.items():
            rows = [int(item["__source_row__"]) for item in items]
            derived = STATS.binary_curve(
                [item.get(label_name) for item in items],
                [finite_number(item.get(score_name), column=score_name, row=int(item["__source_row__"])) for item in items],
                rows, positive_label, curve, bool(config["parameters"].get("compute_auc", False)),
            )
            label = group + (f" (AUC={derived['auc']:.3f})" if derived["auc"] is not None else "")
            ax.plot([point["x"] for point in derived["points"]], [point["y"] for point in derived["points"]], marker="o", markersize=3, label=label)
            marks.extend({"x": point["x"], "y": point["y"], "group": group, "source_rows": point["source_rows"], "threshold": point["threshold"], "derived": derived} for point in derived["points"])
        if group_name or bool(config["parameters"].get("compute_auc", False)): ax.legend(frameon=False)
    else:
        x_name, y_name, group_name = str(panel["x"]), str(panel["y"]), panel.get("group")
        grouped: dict[str, list[tuple[float, float, int]]] = {}
        for record in records:
            row = int(record["__source_row__"]); group = str(record.get(group_name, "All")) if group_name else "All"
            x = finite_number(record.get(x_name), column=x_name, row=row); y = finite_number(record.get(y_name), column=y_name, row=row)
            if not 0 <= x <= 1 or not 0 <= y <= 1: raise ValueError(f"ROC/PR points must lie in [0,1] at source row {row}")
            grouped.setdefault(group, []).append((x, y, row))
        for group, points in grouped.items():
            points.sort();
            if any(right[0] < left[0] for left, right in zip(points, points[1:])): raise ValueError("curve x values must be nondecreasing")
            ax.plot([point[0] for point in points], [point[1] for point in points], marker="o", label=group if group_name else None)
            marks.extend({"source_row": row, "x": x, "y": y, "group": group} for x, y, row in points)
        if group_name: ax.legend(frameon=False)
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.set_xlabel("False Positive Rate" if curve == "roc" else "Recall")
    ax.set_ylabel("True Positive Rate" if curve == "roc" else "Precision")
    ax.grid(color="#dddddd", linewidth=0.7)
    return marks


def render_to_axis(ax, fig, panel: dict, records: list[dict[str, Any]]) -> list[dict]:
    visual_form = panel.get("visual_form")
    if visual_form == "bar-chart": return plot_bar(ax, panel, records)
    if visual_form == "line-chart": return plot_xy(ax, panel, records, scatter=False)
    if visual_form == "scatter-plot": return plot_xy(ax, panel, records, scatter=True)
    if visual_form == "heatmap": return plot_heatmap(ax, fig, panel, records)
    if visual_form == "box-plot": return plot_box(ax, panel, records)
    if visual_form == "violin-plot": return plot_density_or_violin(ax, panel, records, violin=True)
    if visual_form == "histogram": return plot_histogram(ax, panel, records)
    if visual_form == "density-plot": return plot_density_or_violin(ax, panel, records, violin=False)
    if visual_form == "confusion-matrix": return plot_confusion(ax, fig, panel, records)
    if visual_form == "roc-curve": return plot_curve(ax, panel, records, curve="roc")
    if visual_form == "pr-curve": return plot_curve(ax, panel, records, curve="pr")
    raise ValueError(f"unsupported visual form: {visual_form}")


def apply_axis_config(axes: list[Any], panel: dict, marks: list[dict[str, Any]]) -> None:
    config = panel.get("axis") or {}
    if not isinstance(config, dict): raise ValueError("axis must be an object")
    x_scale, y_scale = config.get("x_scale", "linear"), config.get("y_scale", "linear")
    if x_scale not in {"linear", "log", "symlog"} or y_scale not in {"linear", "log", "symlog"}:
        raise ValueError("axis scale must be linear, log, or symlog")
    numeric_x = [float(mark["x"]) for mark in marks if isinstance(mark.get("x"), (int, float))]
    numeric_y = [float(mark["y"]) for mark in marks if isinstance(mark.get("y"), (int, float))]
    if x_scale == "log" and any(value <= 0 for value in numeric_x): raise ValueError("log x axis requires positive values")
    if y_scale == "log":
        if any(value <= 0 for value in numeric_y): raise ValueError("log y axis requires positive values")
        for mark in marks:
            uncertainty = mark.get("uncertainty")
            if uncertainty and float(mark["y"]) - uncertainty["lower"] <= 0:
                raise ValueError("log y axis uncertainty may not cross zero")
    if "x_scale" in config:
        if not numeric_x and x_scale != "linear": raise ValueError("non-linear x scale requires numeric x values")
        if numeric_x:
            for ax in axes: ax.set_xscale(x_scale, **({"linthresh": config.get("x_linthresh", 1.0)} if x_scale == "symlog" else {}))
    if "y_scale" in config:
        if not numeric_y and y_scale != "linear": raise ValueError("non-linear y scale requires numeric y values")
        if numeric_y:
            for ax in axes: ax.set_yscale(y_scale, **({"linthresh": config.get("y_linthresh", 1.0)} if y_scale == "symlog" else {}))
    if config.get("x_limits") is not None:
        for ax in axes: ax.set_xlim(*config["x_limits"])
    if config.get("y_limits") is not None and len(axes) == 1:
        limits = config["y_limits"]
        if panel.get("visual_form") == "bar-chart" and float(limits[0]) != 0 and not str(config.get("baseline_justification", "")).strip():
            raise ValueError("non-zero bar baseline requires baseline_justification")
        axes[0].set_ylim(*limits)
    axis_break = config.get("break")
    if axis_break:
        if panel.get("visual_form") not in {"line-chart", "scatter-plot"}:
            raise ValueError("axis breaks are supported only for line and scatter plots")
        if not str(axis_break.get("justification", "")).strip(): raise ValueError("axis break requires justification")
        omit = axis_break.get("omit")
        if axis_break.get("axis") != "y" or not isinstance(omit, list) or len(omit) != 2 or float(omit[1]) <= float(omit[0]):
            raise ValueError("axis break requires axis='y' and increasing omit bounds")
        low, high = map(float, omit)
        if any(low < value < high for value in numeric_y): raise ValueError("data point lies inside omitted axis interval")
        bottom, top = axes[1], axes[0]
        overall_low = min(numeric_y) if numeric_y else low - 1
        overall_high = max(numeric_y) if numeric_y else high + 1
        bottom.set_ylim(overall_low - max(1e-9, abs(overall_low) * 0.05), low)
        top.set_ylim(high, overall_high + max(1e-9, abs(overall_high) * 0.05))
        top.spines["bottom"].set_visible(False); bottom.spines["top"].set_visible(False)
        top.tick_params(labeltop=False, bottom=False); bottom.xaxis.tick_bottom()
        kwargs = dict(marker=[(-1, -0.5), (1, 0.5)], markersize=8, linestyle="none", color="k", mec="k", mew=1, clip_on=False)
        top.plot([0, 1], [0, 0], transform=top.transAxes, **kwargs); bottom.plot([0, 1], [1, 1], transform=bottom.transAxes, **kwargs)


def provenance_marks(panel: dict, marks: list[dict[str, Any]], subplot_id: str | None = None) -> list[dict[str, Any]]:
    result = []
    for mark in marks:
        item: dict[str, Any] = {"transform": panel.get("transform", "none")}
        if subplot_id is not None: item["subplot"] = subplot_id
        if "source_row" in mark: item["source_row"] = mark["source_row"]
        if "source_rows" in mark: item["source_rows"] = mark["source_rows"]
        for key in ("x", "y", "value", "x2"):
            if key in mark:
                column = panel.get(key)
                item[key] = {"column": column, "value": mark[key]}
                if key in {"y", "value"}: item[key]["unit"] = panel.get("unit")
        if "group" in mark:
            item["group"] = {"column": panel.get("group") or panel.get("x"), "value": mark["group"]}
        if "uncertainty" in mark:
            item["uncertainty"] = mark["uncertainty"]
            if panel.get("error") and mark["uncertainty"]["mode"] == "symmetric-delta":
                item["error"] = {"column": panel["error"], "value": mark["uncertainty"]["values"]["error"], "semantics": "symmetric-absolute"}
        if "derived" in mark: item["derived"] = mark["derived"]
        if "precomputed" in mark: item["precomputed"] = mark["precomputed"]
        if "threshold" in mark: item["threshold"] = mark["threshold"]
        result.append(item)
    return result


def render_panel(panel: dict, input_root: Path, output_dir: Path, formats: Iterable[str]) -> dict:
    plt = require_matplotlib()
    source, records = panel_records(panel, input_root)
    if not records:
        raise ValueError(f"panel {panel.get('id')} data source is empty")
    visual_form = panel.get("visual_form")
    axis_break = (panel.get("axis") or {}).get("break")
    if axis_break:
        fig, raw_axes = plt.subplots(2, 1, figsize=(5.4, 4.8), sharex=True, gridspec_kw={"height_ratios": [1, 1], "hspace": 0.08})
        axes = list(raw_axes)
    else:
        fig, ax = plt.subplots(figsize=(6.0, 4.4) if visual_form in {"heatmap", "confusion-matrix"} else (5.4, 3.7))
        axes = [ax]
    try:
        marks = render_to_axis(axes[0], fig, panel, records)
        for extra in axes[1:]: render_to_axis(extra, fig, panel, records)
        apply_axis_config(axes, panel, marks)
        axes[0].set_title(str(panel.get("title") or panel.get("id")), loc="left", fontsize=11, fontweight="bold", pad=10)
        if axis_break:
            fig.subplots_adjust(left=0.14, right=0.97, top=0.91, bottom=0.12, hspace=0.08)
        else:
            fig.tight_layout()
        stem = f"panel_{str(panel['id']).lower()}"
        outputs = {}
        for fmt in formats:
            target = output_dir / f"{stem}.{fmt}"
            kwargs = {"dpi": 220} if fmt == "png" else {}
            fig.savefig(target, bbox_inches="tight", **kwargs)
            outputs[fmt] = target.name
    finally:
        plt.close(fig)

    return {
        "panel": panel["id"],
        "visual_form": visual_form,
        "source_file": str(source),
        "source_sha256": sha256(source),
        "outputs": outputs,
        "marks": provenance_marks(panel, marks),
        "calculation": panel.get("calculation"), "axis": panel.get("axis"), "uncertainty": panel.get("uncertainty"),
    }


def render_grid_panel(panel: dict, input_root: Path, output_dir: Path, formats: Iterable[str]) -> dict:
    plt = require_matplotlib()
    subplots = panel.get("subplots") or []
    layout = panel.get("layout") or {}
    rows, columns = layout.get("rows"), layout.get("columns")
    if not isinstance(rows, int) or not isinstance(columns, int) or rows < 1 or columns < 1 or rows * columns < len(subplots):
        raise ValueError("data-plot-grid requires positive rows/columns covering every subplot")
    if not subplots:
        raise ValueError("data-plot-grid requires subplots")
    fig, raw_axes = plt.subplots(rows, columns, figsize=(5.2 * columns, 3.7 * rows), squeeze=False, sharex=bool(panel.get("share_x")), sharey=bool(panel.get("share_y")))
    flattened = list(raw_axes.flat)
    provenance = []
    sources = []
    legend_handles, legend_labels = [], []
    try:
        for index, subplot in enumerate(subplots):
            if (subplot.get("axis") or {}).get("break"):
                raise ValueError("axis breaks are not supported inside data-plot-grid")
            source, records = panel_records(subplot, input_root)
            sources.append({"file": str(source), "sha256": sha256(source)})
            ax = flattened[index]
            marks = render_to_axis(ax, fig, subplot, records)
            apply_axis_config([ax], subplot, marks)
            ax.set_title(str(subplot.get("title") or subplot.get("id") or index + 1), loc="left", fontsize=10, fontweight="bold")
            subplot_marks = provenance_marks(subplot, marks, str(subplot.get("id") or index + 1))
            for item in subplot_marks: item["source_file"] = str(source)
            provenance.extend(subplot_marks)
            handles, labels = ax.get_legend_handles_labels()
            for handle, label in zip(handles, labels):
                if label and label not in legend_labels:
                    legend_handles.append(handle); legend_labels.append(label)
            if panel.get("shared_legend") and ax.get_legend() is not None:
                ax.get_legend().remove()
        for ax in flattened[len(subplots):]: ax.set_visible(False)
        if panel.get("shared_legend") and legend_handles:
            fig.legend(legend_handles, legend_labels, loc="upper center", ncol=max(1, len(legend_labels)), frameon=False)
            fig.subplots_adjust(top=0.86)
        else:
            fig.tight_layout()
        stem = f"panel_{str(panel['id']).lower()}"; outputs = {}
        for fmt in formats:
            target = output_dir / f"{stem}.{fmt}"; fig.savefig(target, bbox_inches="tight", **({"dpi": 220} if fmt == "png" else {})); outputs[fmt] = target.name
    finally:
        plt.close(fig)
    return {"panel": panel["id"], "visual_form": "data-plot-grid", "sources": sources, "outputs": outputs, "marks": provenance, "grid": {"rows": rows, "columns": columns, "share_x": bool(panel.get("share_x")), "share_y": bool(panel.get("share_y")), "shared_legend": bool(panel.get("shared_legend"))}}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("plan", type=Path)
    parser.add_argument("--input-root", type=Path)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--panel", action="append", help="Render only the selected panel id; may be repeated")
    parser.add_argument("--formats", default="svg,pdf,png")
    parser.add_argument("--allow-open-questions", action="store_true")
    args = parser.parse_args()

    plan = json.loads(args.plan.read_text(encoding="utf-8"))
    if plan.get("open_questions") and not args.allow_open_questions:
        raise SystemExit("figure plan has unresolved open_questions; resolve them or pass --allow-open-questions for a draft")
    input_root = (args.input_root or Path(plan.get("input_root") or ".")).resolve()
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    formats = tuple(item.strip().lower() for item in args.formats.split(",") if item.strip())
    unsupported = set(formats) - {"svg", "pdf", "png"}
    if unsupported:
        parser.error(f"unsupported formats: {', '.join(sorted(unsupported))}")

    selected = set(args.panel or [])
    panels = [panel for panel in plan.get("panels", []) if panel.get("type") in {"data-plot", "data-plot-grid"} and (not selected or str(panel.get("id")) in selected)]
    if not panels:
        raise SystemExit("no matching data-plot panels found in the plan")

    plt = require_matplotlib()
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "font.size": 9,
        "axes.linewidth": 0.9,
        "svg.fonttype": "none",
        "pdf.fonttype": 42,
    })
    rendered = [render_grid_panel(panel, input_root, output_dir, formats) if panel.get("type") == "data-plot-grid" else render_panel(panel, input_root, output_dir, formats) for panel in panels]
    provenance = {"schema_version": "1.0", "plan": str(args.plan.resolve()), "panels": rendered}
    provenance_path = output_dir / "provenance.json"
    provenance_path.write_text(json.dumps(provenance, ensure_ascii=False, indent=2), encoding="utf-8")
    for panel in panels:
        target_source = output_dir / f"panel_{str(panel['id']).lower()}_source.py"
        if Path(__file__).resolve() != target_source.resolve():
            shutil.copy2(Path(__file__), target_source)
    core_target = output_dir / "statistics_core.py"
    if Path(STATS.__file__).resolve() != core_target.resolve():
        shutil.copy2(Path(STATS.__file__), core_target)
    recipe = {
        "command": "python panel_<id>_source.py <figure-plan.json> --output-dir <output-dir>",
        "input_root": str(input_root),
        "formats": list(formats),
    }
    (output_dir / "render-recipe.json").write_text(json.dumps(recipe, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Rendered {len(rendered)} data panel(s) -> {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
