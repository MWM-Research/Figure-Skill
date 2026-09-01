#!/usr/bin/env python3
"""Run structural, provenance, and deliverable QA on scientific figures."""

from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import json
import math
import re
import struct
import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any


HERE = Path(__file__).resolve().parent
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))

from review_generated_figure import validate_review  # noqa: E402


def load_statistics_core():
    path = HERE / "backends" / "statistics_core.py"
    spec = importlib.util.spec_from_file_location("qa_statistics_core", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load statistics core: {path}")
    module = importlib.util.module_from_spec(spec); spec.loader.exec_module(module); return module


STATS = load_statistics_core()


EDITABLE = {".svg", ".drawio", ".py", ".r", ".ipynb", ".ai", ".eps"}
PLACEHOLDER = re.compile(r"\b(?:todo|tbd|placeholder|lorem ipsum)\b", re.IGNORECASE)


def check(name: str, status: str, **details: Any) -> dict:
    return {"check": name, "status": status, **details}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inspect_svg(path: Path) -> list[dict]:
    checks = []
    try:
        root = ET.parse(path).getroot()
        checks.append(check("svg-xml-valid", "pass", file=str(path)))
        has_size = bool(root.get("viewBox") or (root.get("width") and root.get("height")))
        checks.append(check("svg-has-canvas-size", "pass" if has_size else "fail", file=str(path)))
        text = " ".join(root.itertext())
        checks.append(check("svg-no-placeholders", "fail" if PLACEHOLDER.search(text) else "pass", file=str(path)))
        images = [element for element in root.iter() if element.tag.rsplit("}", 1)[-1] == "image"]
        checks.append(check(
            "svg-not-raster-only",
            "warn" if images else "pass",
            file=str(path),
            detail=f"contains {len(images)} embedded image element(s)" if images else "no embedded raster images",
        ))
    except (ET.ParseError, OSError) as exc:
        checks.append(check("svg-xml-valid", "fail", file=str(path), detail=str(exc)))
    return checks


def inspect_pdf(path: Path) -> list[dict]:
    try:
        data = path.read_bytes()
    except OSError as exc:
        return [check("pdf-readable", "fail", file=str(path), detail=str(exc))]
    valid = len(data) > 100 and data.startswith(b"%PDF-") and b"%%EOF" in data[-2048:]
    return [check("pdf-readable", "pass" if valid else "fail", file=str(path), size_bytes=len(data))]


def inspect_png(path: Path) -> list[dict]:
    try:
        with path.open("rb") as handle:
            header = handle.read(24)
        valid = len(header) == 24 and header[:8] == b"\x89PNG\r\n\x1a\n"
        width, height = struct.unpack(">II", header[16:24]) if valid else (0, 0)
        status = "pass" if valid and width >= 300 and height >= 200 else ("warn" if valid else "fail")
        return [check("png-valid-size", status, file=str(path), width=width, height=height)]
    except OSError as exc:
        return [check("png-valid-size", "fail", file=str(path), detail=str(exc))]


def png_dimensions(path: Path) -> tuple[int, int] | None:
    try:
        with path.open("rb") as handle:
            header = handle.read(24)
        if len(header) != 24 or header[:8] != b"\x89PNG\r\n\x1a\n":
            return None
        return struct.unpack(">II", header[16:24])
    except OSError:
        return None


def equal_value(raw: Any, expected: Any) -> bool:
    try:
        return math.isclose(float(str(raw).replace(",", "")), float(expected), rel_tol=1e-9, abs_tol=1e-12)
    except (TypeError, ValueError):
        return str(raw) == str(expected)


def source_rows(path: Path) -> dict[int, dict[str, Any]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            return {index: row for index, row in enumerate(csv.DictReader(handle, delimiter=delimiter), start=2)}
    if suffix == ".jsonl":
        result = {}
        with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
            for index, line in enumerate(handle, start=1):
                if line.strip():
                    value = json.loads(line)
                    if isinstance(value, dict):
                        result[index] = value
        return result
    if suffix == ".json":
        value = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(value, dict):
            value = next((value[key] for key in ("records", "data", "results", "metrics", "items") if isinstance(value.get(key), list)), [value])
        return {index: item for index, item in enumerate(value, start=1) if isinstance(item, dict)} if isinstance(value, list) else {}
    if suffix == ".xlsx":
        from openpyxl import load_workbook  # type: ignore
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else f"column_{index + 1}" for index, value in enumerate(next(iterator, []))]
        try:
            return {index: dict(zip(headers, values)) for index, values in enumerate(iterator, start=2)}
        finally:
            workbook.close()
    raise ValueError(f"unsupported provenance source format: {suffix}")


def values_close(left: Any, right: Any) -> bool:
    if isinstance(left, dict) and isinstance(right, dict):
        return left.keys() == right.keys() and all(values_close(left[key], right[key]) for key in left)
    if isinstance(left, list) and isinstance(right, list):
        return len(left) == len(right) and all(values_close(a, b) for a, b in zip(left, right))
    return equal_value(left, right)


def plan_panel_index(plan: dict | None) -> dict[tuple[str, str | None], dict]:
    result = {}
    for panel in (plan or {}).get("panels", []):
        panel_id = str(panel.get("id"))
        result[(panel_id, None)] = panel
        for subplot in panel.get("subplots", []):
            result[(panel_id, str(subplot.get("id")))] = subplot
    return result


def recompute_derived(mark: dict, panel: dict, rows: dict[int, dict[str, Any]]) -> tuple[bool, str]:
    derived = mark.get("derived")
    if not isinstance(derived, dict): return True, ""
    operation = derived.get("operation")
    source_row_numbers = [int(value) for value in derived.get("source_rows", mark.get("source_rows", []))]
    selected = [rows[number] for number in source_row_numbers if number in rows]
    if len(selected) != len(source_row_numbers): return False, "derived source row missing"
    try:
        if operation == "box-summary":
            column = str(panel["y"]); expected = STATS.box_summary([float(row[column]) for row in selected], source_row_numbers)
        elif operation == "kde":
            column = str(panel.get("value") or panel.get("y")); expected = STATS.kde([float(row[column]) for row in selected], source_row_numbers, bandwidth=derived["parameters"]["bandwidth"])
        elif operation == "histogram":
            column = str(panel.get("value") or panel.get("x")); expected = STATS.histogram([float(row[column]) for row in selected], source_row_numbers, derived["parameters"])
        elif operation == "confusion-count":
            actual, predicted = str(panel["actual"]), str(panel["predicted"]); expected = STATS.confusion_matrix([str(row[actual]) for row in selected], [str(row[predicted]) for row in selected], source_row_numbers, derived["parameters"]["normalization"])
        elif operation in {"roc", "pr"}:
            label, score = str(panel["label"]), str(panel["score"]); expected = STATS.binary_curve([row[label] for row in selected], [float(row[score]) for row in selected], source_row_numbers, derived["parameters"]["positive_label"], operation, bool(derived["parameters"].get("compute_auc")))
        else:
            return False, f"unsupported derived operation {operation}"
        return values_close(expected, derived), "" if values_close(expected, derived) else f"derived {operation} mismatch"
    except (KeyError, TypeError, ValueError) as exc:
        return False, str(exc)


def verify_data_provenance(path: Path, plan: dict | None = None) -> list[dict]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return [check("provenance-readable", "fail", file=str(path), detail=str(exc))]
    checks = [check("provenance-readable", "pass", file=str(path))]
    panel_lookup = plan_panel_index(plan)
    for panel in data.get("panels", []):
        panel_id = panel.get("panel")
        sources = {str(panel.get("source_file", "")): panel.get("source_sha256")} if panel.get("source_file") else {str(item.get("file")): item.get("sha256") for item in panel.get("sources", [])}
        if not sources:
            checks.append(check("provenance-source-exists", "fail", panel=panel_id, source="")); continue
        source_rows_cache = {}
        for source_text, expected_hash in sources.items():
            source = Path(source_text)
            checks.append(check("provenance-source-exists", "pass" if source.is_file() else "fail", panel=panel_id, source=str(source)))
            if source.is_file():
                checks.append(check("provenance-source-hash", "pass" if expected_hash and sha256(source) == expected_hash else "fail", panel=panel_id, source=str(source)))
                source_rows_cache[str(source)] = source_rows(source)
        try:
            mismatches = []
            derived_mismatches = []
            for mark in panel.get("marks", []):
                source_text = str(mark.get("source_file") or panel.get("source_file"))
                rows = source_rows_cache.get(source_text, {})
                row_numbers = [int(mark["source_row"])] if mark.get("source_row") is not None else [int(value) for value in mark.get("source_rows", [])]
                if any(number not in rows for number in row_numbers): mismatches.append(f"missing source row(s) {row_numbers}")
                if mark.get("source_row") is not None:
                    row_number = int(mark["source_row"]); row = rows.get(row_number, {})
                    for axis in ("x", "y", "group", "value", "error", "x2"):
                        value = mark.get(axis, {})
                        if not value or value.get("column") is None: continue
                        column = value.get("column")
                        if column not in row or not equal_value(row.get(column), value.get("value")): mismatches.append(f"row {row_number} column {column}")
                    uncertainty = mark.get("uncertainty", {})
                    for label, column in uncertainty.get("columns", {}).items():
                        expected = uncertainty.get("values", {}).get(label)
                        if column not in row or not equal_value(row.get(column), expected): mismatches.append(f"row {row_number} column {column}")
                plan_panel = panel_lookup.get((str(panel_id), str(mark.get("subplot")) if mark.get("subplot") is not None else None))
                if mark.get("derived") and plan_panel:
                    valid, detail = recompute_derived(mark, plan_panel, rows)
                    if not valid: derived_mismatches.append(detail)
            checks.append(check(
                "provenance-mark-values", "fail" if mismatches else "pass", panel=panel_id,
                detail=", ".join(mismatches) if mismatches else f"verified {len(panel.get('marks', []))} marks",
            ))
            checks.append(check("provenance-derived-values", "fail" if derived_mismatches else "pass", panel=panel_id, detail=", ".join(dict.fromkeys(derived_mismatches)) if derived_mismatches else "verified"))
        except (OSError, csv.Error, ValueError, TypeError) as exc:
            checks.append(check("provenance-mark-values", "fail", panel=panel_id, detail=str(exc)))
    return checks


def verify_edit_provenance(path: Path) -> list[dict]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        source = Path(str(data.get("source_file", "")))
        valid_source = source.is_file() and data.get("source_sha256") == sha256(source)
        operations = data.get("operations", [])
        applied = bool(operations) and all(item.get("status") == "applied" for item in operations)
        return [
            check("edit-provenance-readable", "pass", file=str(path)),
            check("edit-source-hash", "pass" if valid_source else "fail", source=str(source)),
            check("edit-operations-applied", "pass" if applied else "fail", count=len(operations)),
        ]
    except (OSError, json.JSONDecodeError, TypeError) as exc:
        return [check("edit-provenance-readable", "fail", file=str(path), detail=str(exc))]


def verify_generation_provenance(path: Path) -> list[dict]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        output = Path(str(data.get("output", "")))
        valid_output = output.is_file() and data.get("output_sha256") == sha256(output)
        labeled = data.get("generated_content") is True and data.get("evidence_role") == "illustrative"
        review_required = data.get("human_review_required") is True
        return [
            check("generation-provenance-readable", "pass", file=str(path)),
            check("generation-output-hash", "pass" if valid_output else "fail", output=str(output)),
            check("generated-content-labeled", "pass" if labeled else "fail"),
            check("generation-human-review-required", "pass" if review_required else "fail"),
        ]
    except (OSError, json.JSONDecodeError, TypeError) as exc:
        return [check("generation-provenance-readable", "fail", file=str(path), detail=str(exc))]


def verify_annotation_provenance(path: Path, panel: dict) -> list[dict]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        output = Path(str(data.get("output", "")))
        overlay = Path(str(data.get("overlay_source", "")))
        valid_output = output.is_file() and data.get("output_sha256") == sha256(output)
        valid_overlay = overlay.is_file() and data.get("overlay_source_sha256") == sha256(overlay)
        planned_labels = sorted(str(value) for value in panel.get("visible_labels", []))
        actual_labels = sorted(str(value) for value in data.get("visible_labels", []))
        labels_match = bool(planned_labels) and planned_labels == actual_labels
        canvas = panel.get("canvas", {})
        canvas_match = data.get("canvas") == [canvas.get("width"), canvas.get("height")]
        return [
            check("annotation-provenance-readable", "pass", file=str(path)),
            check("annotation-output-hash", "pass" if valid_output else "fail", output=str(output)),
            check("annotation-source-hash", "pass" if valid_overlay else "fail", source=str(overlay)),
            check("annotation-visible-labels", "pass" if labels_match else "fail", count=len(actual_labels)),
            check("annotation-canvas", "pass" if canvas_match else "fail", value=data.get("canvas")),
        ]
    except (OSError, json.JSONDecodeError, TypeError) as exc:
        return [check("annotation-provenance-readable", "fail", file=str(path), detail=str(exc))]


def verify_hybrid_audit(path: Path) -> list[dict]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        source = Path(str(data.get("source", "")))
        plan_path = Path(str(data.get("plan", "")))
        source_valid = source.is_file() and data.get("source_sha256") == sha256(source)
        plan_valid = plan_path.is_file() and data.get("plan_sha256") == sha256(plan_path)
        checks_pass = bool(data.get("checks")) and all(item.get("status") == "pass" for item in data["checks"])
        return [
            check("hybrid-audit-readable", "pass", file=str(path)),
            check("hybrid-audit-source-hash", "pass" if source_valid else "fail", source=str(source)),
            check("hybrid-audit-plan-hash", "pass" if plan_valid else "fail", plan=str(plan_path)),
            check("hybrid-audit-contract", "pass" if data.get("status") == "pass" and checks_pass else "fail"),
        ]
    except (OSError, json.JSONDecodeError, TypeError) as exc:
        return [check("hybrid-audit-readable", "fail", file=str(path), detail=str(exc))]


def run_qa(target: Path, plan: dict | None) -> dict:
    files = [target] if target.is_file() else sorted(path for path in target.rglob("*") if path.is_file())
    suffixes = {path.suffix.lower() for path in files}
    raster_route = bool(plan and plan.get("route") in {"raster-illustration", "hybrid-composite"})
    technical_checks = [
        check("target-has-files", "pass" if files else "fail"),
        check("files-nonempty", "pass" if files and all(path.stat().st_size > 0 for path in files) else "fail"),
        check(
            "editable-source-present",
            "pass" if raster_route or suffixes & EDITABLE else "fail",
            detail="not required for an explicitly generated raster illustration" if raster_route else "required",
        ),
        check(
            "vector-export-present",
            "pass" if raster_route or suffixes & {".svg", ".pdf", ".eps"} else "fail",
            detail="not required for raster-illustration route" if raster_route else "required",
        ),
        check("preview-present", "pass" if ".png" in suffixes else "warn"),
    ]
    for path in files:
        if path.suffix.lower() == ".svg":
            technical_checks.extend(inspect_svg(path))
        elif path.suffix.lower() == ".pdf":
            technical_checks.extend(inspect_pdf(path))
        elif path.suffix.lower() == ".png":
            technical_checks.extend(inspect_png(path))

    if plan:
        technical_checks.append(check(
            "plan-evidence-guard",
            "pass" if plan.get("constraints", {}).get("forbid_invented_quantitative_claims") is True else "fail",
        ))
        technical_checks.append(check("plan-open-questions-resolved", "pass" if not plan.get("open_questions") else "warn"))
        review_status = plan.get("review_status")
        technical_checks.append(check("plan-reviewed", "pass" if review_status == "approved" else "warn", value=review_status))
        planned = [
            (str(panel.get("id", "")).lower(), ".png" if panel.get("type") in {"raster-illustration", "hybrid-composite"} else ".svg")
            for panel in plan.get("panels", [])
        ]
        missing_panels = [
            panel_id for panel_id, suffix in planned
            if not any(path.name.lower() == f"panel_{panel_id}{suffix}" for path in files)
        ]
        technical_checks.append(check(
            "planned-panels-present", "fail" if missing_panels else "pass",
            detail=f"missing: {', '.join(missing_panels)}" if missing_panels else f"found {len(planned)} panel(s)",
        ))
        has_data_panels = any(panel.get("type") == "data-plot" for panel in plan.get("panels", []))
        if has_data_panels:
            technical_checks.append(check("data-render-source-present", "pass" if ".py" in suffixes else "fail"))
            provenance_files = [path for path in files if path.name in {"data-provenance.json", "provenance.json"}]
            technical_checks.append(check("data-provenance-present", "pass" if provenance_files else "fail"))
            for path in provenance_files:
                technical_checks.extend(verify_data_provenance(path, plan))
        has_edit_panels = any(panel.get("type") == "edit" for panel in plan.get("panels", []))
        if has_edit_panels:
            edit_files = [path for path in files if path.name == "edit-provenance.json"]
            technical_checks.append(check("edit-provenance-present", "pass" if edit_files else "fail"))
            for path in edit_files:
                technical_checks.extend(verify_edit_provenance(path))
        has_raster_panels = any(panel.get("type") in {"raster-illustration", "hybrid-composite"} for panel in plan.get("panels", []))
        if has_raster_panels:
            generation_files = [path for path in files if path.name == "generation-provenance.json"]
            request_files = [path for path in files if path.name == "raster-illustration-request.json"]
            technical_checks.append(check("generation-provenance-present", "pass" if generation_files else "fail"))
            technical_checks.append(check("generation-request-present", "pass" if request_files else "fail"))
            for path in generation_files:
                technical_checks.extend(verify_generation_provenance(path))
            for panel in (panel for panel in plan.get("panels", []) if panel.get("type") in {"raster-illustration", "hybrid-composite"}):
                if panel.get("annotation_spec", {}).get("mode") == "deterministic-overlay":
                    annotation_files = [path for path in files if path.name == "annotation-provenance.json"]
                    technical_checks.append(check(
                        "annotation-provenance-present", "pass" if len(annotation_files) == 1 else "fail",
                        count=len(annotation_files),
                    ))
                    for path in annotation_files:
                        technical_checks.extend(verify_annotation_provenance(path, panel))
                panel_id = str(panel.get("id", "")).lower()
                panel_path = next((path for path in files if path.name.lower() == f"panel_{panel_id}.png"), None)
                canvas = panel.get("canvas", {})
                expected = (canvas.get("width"), canvas.get("height"))
                actual = png_dimensions(panel_path) if panel_path else None
                size_ok = bool(
                    actual and all(isinstance(value, int) and value > 0 for value in expected)
                    and actual == expected
                )
                technical_checks.append(check(
                    "raster-canvas-size-exact", "pass" if size_ok else "fail",
                    panel=panel.get("id"), expected=list(expected), actual=list(actual) if actual else None,
                ))
                final_path = next(
                    (path for path in files if path.name.lower() == "figure.png" and path.parent.name.lower() == "final"),
                    None,
                )
                final_matches = bool(
                    panel_path and final_path and sha256(panel_path) == sha256(final_path)
                )
                technical_checks.append(check(
                    "raster-final-matches-reviewed-panel", "pass" if final_matches else "fail",
                    panel=str(panel_path) if panel_path else None,
                    final=str(final_path) if final_path else None,
                ))
        has_representation_contract = isinstance(plan.get("representation_contract"), dict) or any(
            isinstance(panel.get("representation_contract"), dict) for panel in plan.get("panels", [])
        )
        if has_representation_contract:
            audit_files = [path for path in files if path.name == "hybrid-structure-audit.json"]
            technical_checks.append(check(
                "hybrid-structure-audit-present", "pass" if len(audit_files) == 1 else "fail", count=len(audit_files)
            ))
            for path in audit_files:
                technical_checks.extend(verify_hybrid_audit(path))

    scientific_status = "not-applicable"
    human_review_status = "not-required"
    scientific_checks: list[dict] = []
    if raster_route:
        review_files = [path for path in files if path.name == "scientific-review.json"]
        if len(review_files) == 1:
            try:
                review = json.loads(review_files[0].read_text(encoding="utf-8"))
                validation = validate_review(review)
                scientific_status = validation["scientific_status"]
                human_review_status = validation["human_review_status"]
                scientific_checks.extend(validation["checks"])
            except (OSError, json.JSONDecodeError, TypeError, ValueError) as exc:
                scientific_status = "fail"
                human_review_status = "rejected"
                scientific_checks.append(check("scientific-review-readable", "fail", detail=str(exc)))
        else:
            scientific_status = "pending" if not review_files else "fail"
            human_review_status = "pending" if not review_files else "rejected"
            scientific_checks.append(check(
                "scientific-review-present", "warn" if not review_files else "fail",
                detail="prepare reports/scientific-review.json and complete assessment plus human approval",
            ))

    technical_status = "fail" if any(item["status"] == "fail" for item in technical_checks) else (
        "warn" if any(item["status"] == "warn" for item in technical_checks) else "pass"
    )
    if technical_status == "fail" or scientific_status == "fail" or human_review_status == "rejected":
        status = "fail"
    elif technical_status == "warn" or scientific_status == "pending" or human_review_status == "pending":
        status = "warn"
    else:
        status = "pass"
    return {
        "schema_version": "1.2",
        "status": status,
        "technical_status": technical_status,
        "scientific_status": scientific_status,
        "human_review_status": human_review_status,
        "target": str(target.resolve()),
        "checks": technical_checks + scientific_checks,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("target", type=Path)
    parser.add_argument("--plan", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    if not args.target.exists():
        parser.error(f"target does not exist: {args.target}")
    plan = json.loads(args.plan.read_text(encoding="utf-8")) if args.plan else None
    report = run_qa(args.target, plan)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"QA status: {report['status']} -> {args.output}")
    return 1 if report["status"] == "fail" else 0


if __name__ == "__main__":
    raise SystemExit(main())
